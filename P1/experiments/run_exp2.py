"""
P1 – Experiment 2: Algorithm optimisation comparison (fixed INDP).

(1) Reward curves over episodes (ACO / Greedy / GA / PPO / IPPO)
(2) Varying noise -> accuracy
(3) Varying noise -> energy
(4) Varying node count -> accuracy
(5) Varying node count -> energy

All algorithms train in parallel via multiprocessing.
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from tqdm import tqdm
import torch

from env.config import DEFAULT_NOISE_FACTOR
from P1.mechanisms.indp import INDPMechanism
from P1.nd_env import NeighborDiscoveryEnv
from P1.algorithms.ppo import PPOAgent
from P1.algorithms.ippo import IPPOAgent
from P1.algorithms.ant_colony import AntColonyOptimizer
from P1.algorithms.greedy import GreedyOptimizer
from P1.algorithms.genetic import GeneticOptimizer

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")
os.makedirs(RESULTS_DIR, exist_ok=True)

EPISODES = 2500
EP_LEN = 50
ALGO_NAMES_REWARD = ["PPO", "IPPO", "ACO", "Greedy", "GA"]
ALGO_NAMES_EVAL = ["ACO", "Greedy", "GA", "PPO", "IPPO"]
STYLES = {"ACO": "D--", "Greedy": "v-.", "GA": "x:",
          "PPO": "s-", "IPPO": "o-"}


# ===================================================================
# Per-algorithm training functions (for parallel execution)
# ===================================================================

def _train_ppo(noise_factor, ep_len, episodes):
    env = NeighborDiscoveryEnv(noise_factor=noise_factor, episode_length=ep_len)
    agent = PPOAgent(env.obs_dim, env.n_actions)
    rewards = []
    for ep in range(episodes):
        obs, _ = env.reset()
        ep_reward = 0.0
        for _ in range(ep_len):
            action, lp, val = agent.select_action(obs)
            obs2, r, done, _, info = env.step(np.array([action]))
            agent.store(obs, action, r, lp, val, done)
            obs = obs2
            ep_reward += r
            if done:
                break
        agent.update()
        rewards.append(ep_reward)
    return agent, rewards


def _train_ippo(noise_factor, ep_len, episodes):
    env = NeighborDiscoveryEnv(noise_factor=noise_factor, episode_length=ep_len)
    agent = IPPOAgent(env.obs_dim, env.n_actions)
    n_virtual_agents = min(env.n_agents, 5)
    rewards = []
    for ep in range(episodes):
        obs, _ = env.reset()
        ep_reward = 0.0
        aid = ep % n_virtual_agents
        for _ in range(ep_len):
            action, lp, val = agent.select_action(obs, agent_id=aid)
            obs2, r, done, _, info = env.step(np.array([action]))
            agent.store(aid, obs, action, r, lp, val, done)
            obs = obs2
            ep_reward += r
            if done:
                break
        agent.update()
        rewards.append(ep_reward)
    return agent, rewards


def _train_aco(noise_factor, ep_len, episodes):
    env = NeighborDiscoveryEnv(noise_factor=noise_factor, episode_length=ep_len)
    aco = AntColonyOptimizer()
    mechanism = INDPMechanism()
    rewards = []
    for ep in range(episodes):
        obs, _ = env.reset()
        p, s, t = aco.select_params()
        params = (aco.POWER_LEVELS[p], aco.SCAN_DURATIONS[s],
                  aco.VERIFY_THRESHOLDS[t])
        ep_reward = 0.0
        mechanism.reset()
        for _ in range(ep_len):
            obs, r, done, _, info = env.step(
                np.array([p * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                          + s * len(env.VERIFY_THRESHOLDS) + t]))
            ep_reward += r
            if done:
                break
        aco.update_pheromone([((p, s, t), max(info.get("accuracy", 0.01), 0.01))])
        rewards.append(ep_reward)
    return aco, rewards


def _train_greedy(noise_factor, ep_len, episodes):
    env = NeighborDiscoveryEnv(noise_factor=noise_factor, episode_length=ep_len)
    greedy = GreedyOptimizer()
    rewards = []
    for ep in range(episodes):
        obs, _ = env.reset()
        ep_reward = 0.0
        params = greedy.select_params()
        p_idx = env.POWER_LEVELS.index(min(env.POWER_LEVELS,
                                           key=lambda x: abs(x - params[0])))
        s_idx = env.SCAN_DURATIONS.index(min(env.SCAN_DURATIONS,
                                             key=lambda x: abs(x - params[1])))
        t_idx = env.VERIFY_THRESHOLDS.index(min(env.VERIFY_THRESHOLDS,
                                                 key=lambda x: abs(x - params[2])))
        action = (p_idx * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                  + s_idx * len(env.VERIFY_THRESHOLDS) + t_idx)
        for _ in range(ep_len):
            obs, r, done, _, info = env.step(np.array([action]))
            ep_reward += r
            if done:
                break
        greedy.update(params, info.get("accuracy", 0.0),
                      info.get("energy", 1.0))
        rewards.append(ep_reward)
    return greedy, rewards


def _train_ga(noise_factor, ep_len, episodes):
    env = NeighborDiscoveryEnv(noise_factor=noise_factor, episode_length=ep_len)
    ga = GeneticOptimizer()
    rewards = []
    for ep in range(episodes):
        obs, _ = env.reset()
        idx = ep % ga.pop_size
        params = ga.get_params(idx)
        p_idx = env.POWER_LEVELS.index(min(env.POWER_LEVELS,
                                           key=lambda x: abs(x - params[0])))
        s_idx = env.SCAN_DURATIONS.index(min(env.SCAN_DURATIONS,
                                             key=lambda x: abs(x - params[1])))
        t_idx = env.VERIFY_THRESHOLDS.index(min(env.VERIFY_THRESHOLDS,
                                                 key=lambda x: abs(x - params[2])))
        action = (p_idx * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                  + s_idx * len(env.VERIFY_THRESHOLDS) + t_idx)
        ep_reward = 0.0
        for _ in range(ep_len):
            obs, r, done, _, info = env.step(np.array([action]))
            ep_reward += r
            if done:
                break
        ga.set_fitness(idx, info.get("accuracy", 0.0),
                       info.get("energy", 1.0))
        if (ep + 1) % ga.pop_size == 0:
            ga.evolve()
        rewards.append(ep_reward)
    return ga, rewards


# ===================================================================
# Parallel workers (module-level for multiprocessing spawn)
# ===================================================================

def _init_pool_worker(gpu_frac):
    import matplotlib as _mpl
    _mpl.use("Agg")
    os.environ.setdefault("SDL_VIDEODRIVER", "dummy")
    if gpu_frac > 0 and torch.cuda.is_available():
        torch.cuda.set_per_process_memory_fraction(gpu_frac)


def _pool_worker_train(args):
    """Train a single algorithm, return rewards list."""
    algo_name, noise_factor, ep_len, episodes = args
    if algo_name == "PPO":
        _, rewards = _train_ppo(noise_factor, ep_len, episodes)
    elif algo_name == "IPPO":
        _, rewards = _train_ippo(noise_factor, ep_len, episodes)
    elif algo_name == "ACO":
        _, rewards = _train_aco(noise_factor, ep_len, episodes)
    elif algo_name == "Greedy":
        _, rewards = _train_greedy(noise_factor, ep_len, episodes)
    elif algo_name == "GA":
        _, rewards = _train_ga(noise_factor, ep_len, episodes)
    else:
        rewards = [0.0] * episodes
    return {"algo": algo_name, "rewards": [float(r) for r in rewards]}


def _pool_worker_eval(args):
    """Train + evaluate an algorithm, return (accuracy, energy)."""
    algo_name, noise_factor, ep_len, episodes, n_trials = args
    if algo_name == "PPO":
        agent, _ = _train_ppo(noise_factor, ep_len, episodes)
    elif algo_name == "IPPO":
        agent, _ = _train_ippo(noise_factor, ep_len, episodes)
    elif algo_name == "ACO":
        agent, _ = _train_aco(noise_factor, ep_len, episodes)
    elif algo_name == "Greedy":
        agent, _ = _train_greedy(noise_factor, ep_len, episodes)
    elif algo_name == "GA":
        agent, _ = _train_ga(noise_factor, ep_len, episodes)
    else:
        return {"accuracy": 0.0, "energy": 0.0}

    env = NeighborDiscoveryEnv(noise_factor=noise_factor, episode_length=ep_len)
    accs, energies = [], []
    for _ in range(n_trials):
        obs, _ = env.reset()
        for _ in range(ep_len):
            if algo_name == "IPPO":
                a, _, _ = agent.select_action(obs, agent_id=0)
            elif algo_name == "PPO":
                a, _, _ = agent.select_action(obs)
            elif algo_name == "ACO":
                params = agent.get_action_values()
                p_idx = env.POWER_LEVELS.index(min(env.POWER_LEVELS,
                    key=lambda x: abs(x - params[0])))
                s_idx = env.SCAN_DURATIONS.index(min(env.SCAN_DURATIONS,
                    key=lambda x: abs(x - params[1])))
                t_idx = env.VERIFY_THRESHOLDS.index(min(env.VERIFY_THRESHOLDS,
                    key=lambda x: abs(x - params[2])))
                a = (p_idx * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                     + s_idx * len(env.VERIFY_THRESHOLDS) + t_idx)
            elif algo_name == "Greedy":
                params = agent.select_params()
                p_idx = env.POWER_LEVELS.index(min(env.POWER_LEVELS,
                    key=lambda x: abs(x - params[0])))
                s_idx = env.SCAN_DURATIONS.index(min(env.SCAN_DURATIONS,
                    key=lambda x: abs(x - params[1])))
                t_idx = env.VERIFY_THRESHOLDS.index(min(env.VERIFY_THRESHOLDS,
                    key=lambda x: abs(x - params[2])))
                a = (p_idx * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                     + s_idx * len(env.VERIFY_THRESHOLDS) + t_idx)
            elif algo_name == "GA":
                params = agent.get_best_params()
                p_idx = env.POWER_LEVELS.index(min(env.POWER_LEVELS,
                    key=lambda x: abs(x - params[0])))
                s_idx = env.SCAN_DURATIONS.index(min(env.SCAN_DURATIONS,
                    key=lambda x: abs(x - params[1])))
                t_idx = env.VERIFY_THRESHOLDS.index(min(env.VERIFY_THRESHOLDS,
                    key=lambda x: abs(x - params[2])))
                a = (p_idx * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                     + s_idx * len(env.VERIFY_THRESHOLDS) + t_idx)
            else:
                a = 0
            obs, r, done, _, info = env.step(np.array([a]))
            if done:
                break
        accs.append(info.get("accuracy", 0.0))
        energies.append(info.get("energy", 0.0))
    return {"accuracy": float(np.mean(accs)), "energy": float(np.mean(energies))}


def _eval_agent_on_config(agent, algo_name, eval_noise, ep_len, n_trials, node_counts):
    """Evaluate a pre-trained agent on a specific config (no retraining)."""
    env = NeighborDiscoveryEnv(node_counts=node_counts,
                               noise_factor=eval_noise,
                               episode_length=ep_len)
    accs, energies = [], []
    for _ in range(n_trials):
        obs, _ = env.reset()
        for _ in range(ep_len):
            if algo_name == "IPPO":
                a, _, _ = agent.select_action(obs, agent_id=0)
            elif algo_name == "PPO":
                a, _, _ = agent.select_action(obs)
            else:
                if algo_name == "ACO":
                    params = agent.get_action_values()
                elif algo_name == "Greedy":
                    params = agent.select_params()
                elif algo_name == "GA":
                    params = agent.get_best_params()
                else:
                    params = (0.6, 0.1, 0.5)
                p_idx = env.POWER_LEVELS.index(min(env.POWER_LEVELS,
                    key=lambda x: abs(x - params[0])))
                s_idx = env.SCAN_DURATIONS.index(min(env.SCAN_DURATIONS,
                    key=lambda x: abs(x - params[1])))
                t_idx = env.VERIFY_THRESHOLDS.index(min(env.VERIFY_THRESHOLDS,
                    key=lambda x: abs(x - params[2])))
                a = (p_idx * len(env.SCAN_DURATIONS) * len(env.VERIFY_THRESHOLDS)
                     + s_idx * len(env.VERIFY_THRESHOLDS) + t_idx)
            obs, r, done, _, info = env.step(np.array([a]))
            if done:
                break
        accs.append(info.get("accuracy", 0.0))
        energies.append(info.get("energy", 0.0))
    return {"accuracy": float(np.mean(accs)), "energy": float(np.mean(energies))}


def _pool_worker_train_and_eval_configs(args):
    """Train once, then evaluate on multiple configs."""
    (algo_name, train_noise, eval_configs, ep_len, episodes, n_trials) = args

    if algo_name == "PPO":
        agent, _ = _train_ppo(train_noise, ep_len, episodes)
    elif algo_name == "IPPO":
        agent, _ = _train_ippo(train_noise, ep_len, episodes)
    elif algo_name == "ACO":
        agent, _ = _train_aco(train_noise, ep_len, episodes)
    elif algo_name == "Greedy":
        agent, _ = _train_greedy(train_noise, ep_len, episodes)
    elif algo_name == "GA":
        agent, _ = _train_ga(train_noise, ep_len, episodes)
    else:
        return [{"accuracy": 0.0, "energy": 0.0}] * len(eval_configs)

    results = []
    for eval_noise, node_counts in eval_configs:
        r = _eval_agent_on_config(agent, algo_name, eval_noise, ep_len,
                                   n_trials, node_counts)
        results.append(r)
    return results


def _pool_worker_eval_config(args):
    """Legacy wrapper: train + evaluate on single config."""
    (algo_name, train_noise, eval_noise, ep_len, episodes,
     n_trials, node_counts) = args
    result = _pool_worker_train_and_eval_configs(
        (algo_name, train_noise, [(eval_noise, node_counts)],
         ep_len, episodes, n_trials))
    return result[0]


# ===================================================================
# xlsx saving
# ===================================================================

def _save_reward_xlsx(algo_rewards, episodes):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  [WARN] openpyxl not installed, skipping xlsx save")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Reward_Curves"
    ws.append(["Episode"] + list(algo_rewards.keys()))
    for ep in range(episodes):
        row = [ep + 1]
        for algo, rewards in algo_rewards.items():
            row.append(rewards[ep] if ep < len(rewards) else None)
        ws.append(row)
    wb.save(os.path.join(RESULTS_DIR, "exp2_1_reward_data.xlsx"))
    print("  Saved exp2_1_reward_data.xlsx")


def _save_eval_xlsx(noise_factors, acc_results, eng_results,
                    totals, acc_results2, eng_results2, algo_names):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  [WARN] openpyxl not installed, skipping xlsx save")
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Exp2_2_Noise_Accuracy"
    ws1.append(["Noise_Factor"] + algo_names)
    for i, nf in enumerate(noise_factors):
        row = [nf]
        for a in algo_names:
            row.append(acc_results[a][i])
        ws1.append(row)

    ws2 = wb.create_sheet("Exp2_3_Noise_Energy")
    ws2.append(["Noise_Factor"] + algo_names)
    for i, nf in enumerate(noise_factors):
        row = [nf]
        for a in algo_names:
            row.append(eng_results[a][i])
        ws2.append(row)

    ws3 = wb.create_sheet("Exp2_4_Nodes_Accuracy")
    ws3.append(["Total_Nodes"] + algo_names)
    for i, t in enumerate(totals):
        row = [t]
        for a in algo_names:
            row.append(acc_results2[a][i])
        ws3.append(row)

    ws4 = wb.create_sheet("Exp2_5_Nodes_Energy")
    ws4.append(["Total_Nodes"] + algo_names)
    for i, t in enumerate(totals):
        row = [t]
        for a in algo_names:
            row.append(eng_results2[a][i])
        ws4.append(row)

    wb.save(os.path.join(RESULTS_DIR, "exp2_eval_data.xlsx"))
    print("  Saved exp2_eval_data.xlsx")


# ===================================================================
# Main experiment functions
# ===================================================================

def run_exp2_1(episodes=EPISODES, ep_len=EP_LEN, noise_factor=2.0,
               max_workers=5, gpu_frac=0.15,
               save_xlsx=True, save_plots=True):
    """Train all algorithms in parallel; return reward curves."""
    print(f"\n=== Exp 2-1: Reward curves ({episodes} episodes, "
          f"{max_workers} workers) ===")

    algo_names = ALGO_NAMES_REWARD
    tasks = [(algo, noise_factor, ep_len, episodes) for algo in algo_names]

    if max_workers > 1:
        from concurrent.futures import ProcessPoolExecutor, as_completed
        import multiprocessing
        ctx = multiprocessing.get_context("spawn")
        with ProcessPoolExecutor(
            max_workers=max_workers, mp_context=ctx,
            initializer=_init_pool_worker, initargs=(gpu_frac,)
        ) as pool:
            fmap = {}
            for i, t in enumerate(tasks):
                fmap[pool.submit(_pool_worker_train, t)] = i
            results = [None] * len(tasks)
            for f in tqdm(as_completed(fmap), total=len(tasks),
                          desc="Training"):
                results[fmap[f]] = f.result()
    else:
        results = []
        for t in tqdm(tasks, desc="Training"):
            results.append(_pool_worker_train(t))

    algo_rewards = {}
    for r in results:
        algo_rewards[r["algo"]] = r["rewards"]

    if save_plots:
        window = 50
        def smooth(x):
            return np.convolve(x, np.ones(window)/window, mode="valid")

        plt.figure(figsize=(10, 6))
        line_styles = {"PPO": "-", "IPPO": "-", "ACO": "--",
                       "Greedy": "-.", "GA": ":"}
        for name in algo_names:
            data = algo_rewards[name]
            plt.plot(smooth(data), line_styles.get(name, "-"),
                     label=name, linewidth=1.5)
        plt.xlabel("Episode", fontsize=12)
        plt.ylabel("Episode Reward", fontsize=12)
        plt.title("Exp 2-1: Reward Curves (INDP + Algorithms)", fontsize=13)
        plt.legend(fontsize=11)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(os.path.join(RESULTS_DIR, "exp2_1_reward_curves.png"),
                    dpi=200)
        plt.close()
        print("  Saved exp2_1_reward_curves.png")

    if save_xlsx:
        _save_reward_xlsx(algo_rewards, episodes)

    return algo_rewards


def run_exp2_2_to_5(episodes=EPISODES, ep_len=EP_LEN,
                    noise_factors=None, scale_factors=None,
                    train_noise=2.0, n_trials=3,
                    max_workers=5, gpu_frac=0.15,
                    save_xlsx=True, save_plots=True):
    """Evaluate all algorithms under varying noise / node count."""
    if noise_factors is None:
        noise_factors = [0.0, 1.0, 2.0, 4.0, 7.0, 10.0]
    if scale_factors is None:
        scale_factors = [0.5, 0.75, 1.0, 1.5, 2.0]

    algo_names = list(ALGO_NAMES_EVAL)
    base = {"satellite": 3, "uav": 6, "ship": 10, "buoy": 20,
            "base_station": 3}

    # -- Exp 2-2 & 2-3: varying noise (train once per algo, eval on all noise levels) --
    print(f"\n=== Exp 2-2/2-3: Varying noise ({max_workers} workers) ===")
    acc_results = {n: [] for n in algo_names}
    eng_results = {n: [] for n in algo_names}

    noise_configs = [(nf, None) for nf in noise_factors]
    tasks_noise = [(algo, train_noise, noise_configs, ep_len, episodes, n_trials)
                   for algo in algo_names]

    if max_workers > 1:
        from concurrent.futures import ProcessPoolExecutor, as_completed
        import multiprocessing
        ctx = multiprocessing.get_context("spawn")
        with ProcessPoolExecutor(
            max_workers=max_workers, mp_context=ctx,
            initializer=_init_pool_worker, initargs=(gpu_frac,)
        ) as pool:
            fmap = {}
            for i, t in enumerate(tasks_noise):
                fmap[pool.submit(_pool_worker_train_and_eval_configs, t)] = i
            algo_noise_results = [None] * len(tasks_noise)
            for f in tqdm(as_completed(fmap), total=len(tasks_noise),
                          desc="Noise eval"):
                algo_noise_results[fmap[f]] = f.result()
    else:
        algo_noise_results = []
        for t in tqdm(tasks_noise, desc="Noise eval"):
            algo_noise_results.append(_pool_worker_train_and_eval_configs(t))

    for ai, algo in enumerate(algo_names):
        for ni in range(len(noise_factors)):
            acc_results[algo].append(algo_noise_results[ai][ni]["accuracy"])
            eng_results[algo].append(algo_noise_results[ai][ni]["energy"])

    if save_plots:
        for metric_name, y_data, ylabel, fname in [
            ("accuracy", acc_results, "Discovery Accuracy",
             "exp2_2_noise_accuracy.png"),
            ("energy", eng_results, "Total Energy Consumption (J)",
             "exp2_3_noise_energy.png"),
        ]:
            plt.figure(figsize=(8, 5))
            for name in algo_names:
                plt.plot(noise_factors, y_data[name], STYLES[name],
                         label=name, linewidth=2, markersize=7)
            plt.xlabel("Environmental Noise Factor", fontsize=12)
            plt.ylabel(ylabel, fontsize=12)
            plt.title(f"Exp 2-{'2' if 'accuracy' in fname else '3'}: "
                      f"{ylabel} vs. Noise", fontsize=13)
            plt.legend(fontsize=11)
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.savefig(os.path.join(RESULTS_DIR, fname), dpi=200)
            plt.close()
            print(f"  Saved {fname}")

    # -- Exp 2-4 & 2-5: varying node count (train once per algo) --
    print(f"\n=== Exp 2-4/2-5: Varying node count ({max_workers} workers) ===")
    acc_results2 = {n: [] for n in algo_names}
    eng_results2 = {n: [] for n in algo_names}
    totals = []

    node_configs = []
    for sf in scale_factors:
        counts = {k: max(1, int(v * sf)) for k, v in base.items()}
        totals.append(sum(counts.values()))
        node_configs.append((train_noise, counts))

    tasks_nodes = [(algo, train_noise, node_configs, ep_len, episodes, n_trials)
                   for algo in algo_names]

    if max_workers > 1:
        from concurrent.futures import ProcessPoolExecutor, as_completed
        import multiprocessing
        ctx2 = multiprocessing.get_context("spawn")
        with ProcessPoolExecutor(
            max_workers=max_workers, mp_context=ctx2,
            initializer=_init_pool_worker, initargs=(gpu_frac,)
        ) as pool:
            fmap = {}
            for i, t in enumerate(tasks_nodes):
                fmap[pool.submit(_pool_worker_train_and_eval_configs, t)] = i
            algo_node_results = [None] * len(tasks_nodes)
            for f in tqdm(as_completed(fmap), total=len(tasks_nodes),
                          desc="Node eval"):
                algo_node_results[fmap[f]] = f.result()
    else:
        algo_node_results = []
        for t in tqdm(tasks_nodes, desc="Node eval"):
            algo_node_results.append(_pool_worker_train_and_eval_configs(t))

    for ai, algo in enumerate(algo_names):
        for ni in range(len(scale_factors)):
            acc_results2[algo].append(algo_node_results[ai][ni]["accuracy"])
            eng_results2[algo].append(algo_node_results[ai][ni]["energy"])

    if save_plots:
        for metric_name, y_data, ylabel, fname in [
            ("accuracy", acc_results2, "Discovery Accuracy",
             "exp2_4_nodes_accuracy.png"),
            ("energy", eng_results2, "Total Energy Consumption (J)",
             "exp2_5_nodes_energy.png"),
        ]:
            plt.figure(figsize=(8, 5))
            for name in algo_names:
                plt.plot(totals, y_data[name], STYLES[name],
                         label=name, linewidth=2, markersize=7)
            plt.xlabel("Total Number of Nodes", fontsize=12)
            plt.ylabel(ylabel, fontsize=12)
            plt.title(f"Exp 2-{'4' if 'accuracy' in fname else '5'}: "
                      f"{ylabel} vs. Node Count", fontsize=13)
            plt.legend(fontsize=11)
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.savefig(os.path.join(RESULTS_DIR, fname), dpi=200)
            plt.close()
            print(f"  Saved {fname}")

    if save_xlsx:
        _save_eval_xlsx(noise_factors, acc_results, eng_results,
                        totals, acc_results2, eng_results2, algo_names)

    return acc_results, eng_results, acc_results2, eng_results2


if __name__ == "__main__":
    run_exp2_1()
    run_exp2_2_to_5()
    print("\nAll Exp-2 plots saved to P1/results/")
